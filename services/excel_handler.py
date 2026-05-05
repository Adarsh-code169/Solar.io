"""
Excel handler for Solar Load Calculator.

Reads a predefined Excel template, identifies INPUT cells (non-formula cells),
fills them with extracted bill data, and generates a completed output file.

CRITICAL: Never overwrites formula cells — only populates designated input cells.
"""

import logging
from datetime import datetime
from pathlib import Path
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import TEMPLATE_FILE, OUTPUT_DIR, SOLAR_DEFAULTS

logger = logging.getLogger(__name__)

# ── Cell Mapping ───────────────────────────────────────────────────────
# Maps extracted JSON field names → Excel cell addresses on "Bill Data" sheet
FIELD_TO_CELL = {
    "consumer_name":    "B3",
    "consumer_number":  "B4",
    "billing_period":   "B5",
    "units_consumed":   "B6",
    "sanctioned_load":  "B7",
    "tariff_category":  "B8",
    "total_bill_amount": "B9",
    "electricity_rate": "B10",
}


def create_template():
    """
    Create the Solar Calculator Excel template with input cells and formula cells.
    This generates the template file used by the application.
    """
    wb = Workbook()

    # ── Styles ─────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, size=11, color="1B5E20")
    label_font = Font(name="Calibri", size=11, color="333333")
    input_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light yellow
    formula_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Light green
    result_font = Font(name="Calibri", bold=True, size=12, color="1B5E20")
    currency_format = '₹#,##0.00'
    percent_format = '0.0%'
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ══════════════════════════════════════════════════════════════════
    # SHEET 1: Bill Data (INPUT SHEET)
    # ══════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Bill Data"
    ws1.sheet_properties.tabColor = "FFC107"

    # Column widths
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 15

    # Header
    ws1.merge_cells("A1:B1")
    ws1["A1"] = "⚡ ELECTRICITY BILL DATA"
    ws1["A1"].font = header_font
    ws1["A1"].fill = header_fill
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["B1"].fill = header_fill
    ws1.row_dimensions[1].height = 35

    # Subtitle
    ws1["A2"] = "Field"
    ws1["B2"] = "Value"
    ws1["C2"] = "Type"
    for col in ["A", "B", "C"]:
        ws1[f"{col}2"].font = Font(name="Calibri", bold=True, size=10, color="666666")
        ws1[f"{col}2"].alignment = Alignment(horizontal="center")

    # Input fields — these are the cells that get populated
    input_fields = [
        ("A3", "Consumer Name",       "B3", "", "C3", "INPUT"),
        ("A4", "Consumer Number",      "B4", "", "C4", "INPUT"),
        ("A5", "Billing Period",       "B5", "", "C5", "INPUT"),
        ("A6", "Units Consumed (kWh)", "B6", 0,  "C6", "INPUT"),
        ("A7", "Sanctioned Load (kW)", "B7", 0,  "C7", "INPUT"),
        ("A8", "Tariff Category",      "B8", "", "C8", "INPUT"),
        ("A9", "Total Bill Amount (₹)","B9", 0,  "C9", "INPUT"),
        ("A10", "Electricity Rate (₹/kWh)", "B10", 0, "C10", "INPUT"),
    ]

    for label_cell, label, value_cell, default, type_cell, cell_type in input_fields:
        ws1[label_cell] = label
        ws1[label_cell].font = label_font
        ws1[label_cell].border = thin_border

        ws1[value_cell] = default
        ws1[value_cell].fill = input_fill
        ws1[value_cell].border = thin_border
        ws1[value_cell].alignment = Alignment(horizontal="center")
        if "₹" in label and isinstance(default, (int, float)):
            ws1[value_cell].number_format = currency_format

        ws1[type_cell] = cell_type
        ws1[type_cell].font = Font(name="Calibri", size=9, color="999999", italic=True)
        ws1[type_cell].alignment = Alignment(horizontal="center")

    # ══════════════════════════════════════════════════════════════════
    # SHEET 2: Solar Calculation (FORMULA SHEET)
    # ══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Solar Calculation")
    ws2.sheet_properties.tabColor = "4CAF50"

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 15

    # Header
    ws2.merge_cells("A1:B1")
    ws2["A1"] = "☀️ SOLAR SYSTEM RECOMMENDATION"
    ws2["A1"].font = header_font
    ws2["A1"].fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    ws2["B1"].fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    ws2.row_dimensions[1].height = 35

    # ── System Parameters (configurable inputs) ───────────────────
    ws2["A2"] = "System Parameters"
    ws2["A2"].font = subheader_font

    params = [
        ("A3", "Peak Sun Hours (hrs/day)",    "B3", SOLAR_DEFAULTS["peak_sun_hours"],        "INPUT"),
        ("A4", "Performance Ratio (0–1)",     "B4", SOLAR_DEFAULTS["performance_ratio"],     "INPUT"),
        ("A5", "Growth Buffer (1 = no buffer)","B5", SOLAR_DEFAULTS["growth_buffer"],        "INPUT"),
        ("A6", "Panel Wattage (W)",           "B6", SOLAR_DEFAULTS["panel_wattage"],         "INPUT"),
        ("A7", "Cost per kW (₹)",             "B7", SOLAR_DEFAULTS["cost_per_kw"],           "INPUT"),
        ("A8", "System Life (years)",         "B8", SOLAR_DEFAULTS["system_life_years"],     "INPUT"),
        ("A9", "Annual Tariff Increase (%)",  "B9", SOLAR_DEFAULTS["annual_tariff_increase"],"INPUT"),
    ]

    for label_cell, label, value_cell, value, cell_type in params:
        ws2[label_cell] = label
        ws2[label_cell].font = label_font
        ws2[label_cell].border = thin_border
        ws2[value_cell] = value
        ws2[value_cell].fill = input_fill
        ws2[value_cell].border = thin_border
        ws2[value_cell].alignment = Alignment(horizontal="center")
        if "₹" in label:
            ws2[value_cell].number_format = currency_format
        if "%" in label:
            ws2[value_cell].number_format = percent_format

    # ── Calculated Results (FORMULAS — DO NOT OVERWRITE) ──────────
    ws2["A9"] = "Solar System Sizing"
    ws2["A9"].font = subheader_font

    ws2["C9"] = "Type"
    ws2["C9"].font = Font(name="Calibri", bold=True, size=10, color="666666")

    # Row 11 onwards (params now occupy rows 3–9, so formulas start at 11)
    # System size formula:
    #   daily_kWh / (peak_sun_hours × performance_ratio) × growth_buffer → round up
    #   This matches pdf_generator.py and script.js exactly.
    formulas = [
        ("A11", "Daily Consumption (kWh/day)",   "B11", "='Bill Data'!B6/30",                                 "FORMULA"),
        ("A12", "Recommended System Size (kW)",   "B12", "=ROUNDUP((B11/(B3*B4))*B5, 0)",                     "FORMULA"),
        ("A13", "Number of Panels (400W)",        "B13", "=ROUNDUP(B12*1000/B6, 0)",                          "FORMULA"),
        ("A14", "Roof Area Required (sq ft)",     "B14", "=B13*20",                                           "FORMULA"),
        ("A15", "Annual Generation (kWh)",        "B15", "=ROUND(B12*B3*B4*365, 0)",                         "FORMULA"),
    ]

    for label_cell, label, value_cell, formula, cell_type in formulas:
        ws2[label_cell] = label
        ws2[label_cell].font = label_font
        ws2[label_cell].border = thin_border
        ws2[value_cell] = formula
        ws2[value_cell].fill = formula_fill
        ws2[value_cell].border = thin_border
        ws2[value_cell].alignment = Alignment(horizontal="center")
        ws2[value_cell].font = result_font

        # Type indicator
        row_num = label_cell[1:]
        ws2[f"C{row_num}"] = cell_type
        ws2[f"C{row_num}"].font = Font(name="Calibri", size=9, color="999999", italic=True)
        ws2[f"C{row_num}"].alignment = Alignment(horizontal="center")

    # ── Financial Analysis (row refs updated: B12=kW, B15=annual gen, B7=cost/kW, B8=life, B9=tariff increase)
    ws2["A17"] = "Financial Analysis"
    ws2["A17"].font = subheader_font

    financial_formulas = [
        ("A18", "Annual Savings (₹)",              "B18", "=ROUND('Bill Data'!B10*B15, 0)",                                  currency_format, "FORMULA"),
        ("A19", "System Cost (₹)",                 "B19", "=ROUND(B12*B7, 0)",                                               currency_format, "FORMULA"),
        ("A20", "Govt Subsidy (₹) — up to 3kW",   "B20", '=IF(B12<=2, 30000*B12, IF(B12<=3, 30000*2+18000*(B12-2), 30000*2+18000*1))', currency_format, "FORMULA"),
        ("A21", "Net Investment (₹)",              "B21", "=B19-B20",                                                        currency_format, "FORMULA"),
        ("A22", "Simple Payback Period (years)",   "B22", "=ROUND(B21/B18, 1)",                                              "0.0",           "FORMULA"),
        ("A23", "25-Year Savings (₹)",             "B23", "=ROUND(B18*((1-(1+B9)^B8)/(-B9)), 0)",                            currency_format, "FORMULA"),
        ("A24", "Return on Investment (%)",        "B24", "=ROUND((B23-B21)/B21*100, 1)",                                    "0.0\"%\"",      "FORMULA"),
        ("A25", "CO₂ Offset (tonnes/year)",        "B25", f"=ROUND(B15*{SOLAR_DEFAULTS['co2_per_kwh']}/1000, 1)",            "0.0",           "FORMULA"),
    ]

    for label_cell, label, value_cell, formula, num_format, cell_type in financial_formulas:
        ws2[label_cell] = label
        ws2[label_cell].font = label_font
        ws2[label_cell].border = thin_border
        ws2[value_cell] = formula
        ws2[value_cell].fill = formula_fill
        ws2[value_cell].border = thin_border
        ws2[value_cell].alignment = Alignment(horizontal="center")
        ws2[value_cell].font = result_font
        ws2[value_cell].number_format = num_format

        row_num = label_cell[1:]
        ws2[f"C{row_num}"] = cell_type
        ws2[f"C{row_num}"].font = Font(name="Calibri", size=9, color="999999", italic=True)
        ws2[f"C{row_num}"].alignment = Alignment(horizontal="center")

    # Save template
    TEMPLATE_FILE.parent.mkdir(exist_ok=True)
    wb.save(str(TEMPLATE_FILE))
    logger.info(f"Template created: {TEMPLATE_FILE}")
    return str(TEMPLATE_FILE)


def populate_template(extracted_data: dict) -> str:
    """
    Load the Excel template, populate INPUT cells with extracted bill data,
    and save as a new output file. Formulas are preserved.

    Args:
        extracted_data: Dictionary of extracted bill fields

    Returns:
        str: Path to the generated output Excel file

    Raises:
        FileNotFoundError: If template file doesn't exist
        ValueError: If extracted data is empty
    """
    if not TEMPLATE_FILE.exists():
        logger.info("Template not found, creating it...")
        create_template()

    if not extracted_data:
        raise ValueError("No extracted data provided")

    logger.info(f"Loading template: {TEMPLATE_FILE}")

    # Load workbook — data_only=False preserves formulas
    wb = load_workbook(str(TEMPLATE_FILE), data_only=False)

    # ── Populate Bill Data sheet ───────────────────────────────────
    ws = wb["Bill Data"]

    for field_name, cell_address in FIELD_TO_CELL.items():
        value = extracted_data.get(field_name)
        if value is not None:
            # Check that we're not overwriting a formula
            current = ws[cell_address].value
            if isinstance(current, str) and current.startswith("="):
                logger.warning(f"SKIPPING {cell_address}: contains formula '{current}'")
                continue

            ws[cell_address] = value
            logger.info(f"SET {cell_address} ({field_name}) = {value}")
        else:
            logger.warning(f"SKIP {cell_address} ({field_name}): value is None")

    # ── Generate output filename ───────────────────────────────────
    consumer = extracted_data.get("consumer_name", "unknown")
    consumer_clean = "".join(c for c in consumer if c.isalnum() or c == " ")[:30].strip().replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"solar_report_{consumer_clean}_{timestamp}.xlsx"
    output_path = OUTPUT_DIR / output_filename

    wb.save(str(output_path))
    logger.info(f"Output saved: {output_path}")

    return str(output_path)


def get_template_info() -> dict:
    """
    Return information about the Excel template structure for debugging.
    """
    if not TEMPLATE_FILE.exists():
        create_template()

    wb = load_workbook(str(TEMPLATE_FILE), data_only=False)

    info = {"sheets": []}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
            for cell in row:
                if cell.value is not None:
                    cell_type = "FORMULA" if isinstance(cell.value, str) and str(cell.value).startswith("=") else "VALUE"
                    cells.append({
                        "cell": cell.coordinate,
                        "value": str(cell.value)[:100],
                        "type": cell_type,
                    })
        info["sheets"].append({"name": sheet_name, "cells": cells})

    return info


# Create template on module import if it doesn't exist
if not TEMPLATE_FILE.exists():
    create_template()
