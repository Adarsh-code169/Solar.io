"""
test_sample.py — Test the full pipeline without a Gemini API key.

Creates sample MSEDCL bill data and generates the Excel output directly.
Run: python test_sample.py
"""

import sys
import json
from pathlib import Path

# Ensure project root is on path
sys.path.insert(0, str(Path(__file__).parent))

# ── Sample Bill Data (simulates what Gemini would extract) ─────────────
SAMPLE_BILL_DATA = {
    "consumer_name": "Rajesh Kumar Sharma",
    "consumer_number": "610230045678",
    "meter_number": "M7823456",
    "billing_period": "01-Jan-2025 to 31-Jan-2025",
    "units_consumed": 320,
    "sanctioned_load": 5.0,
    "tariff_category": "LT-I (Residential)",
    "total_bill_amount": 2854.50,
    "electricity_rate": 8.92,
    "supply_type": "Single Phase",
    "due_date": "15-Feb-2025",
    "previous_reading": 12450,
    "current_reading": 12770,
    "additional_info": "MSEDCL Pune Urban Division - Fuel Adjustment Charges included"
}


def run_test():
    print("=" * 60)
    print("  Solar Load Calculator — Pipeline Test")
    print("=" * 60)

    # Step 1: Show extracted data
    print("\n[1/3] Sample Bill Data (simulates AI extraction):")
    print(json.dumps(SAMPLE_BILL_DATA, indent=2))

    # Step 2: Generate Excel
    print("\n[2/3] Generating Excel report...")
    try:
        from services.excel_handler import populate_template, create_template
        from config import TEMPLATE_FILE

        if not TEMPLATE_FILE.exists():
            print("    Creating Excel template...")
            create_template()
            print("    ✓ Template created")

        output_path = populate_template(SAMPLE_BILL_DATA)
        print(f"    ✓ Excel generated: {output_path}")

    except Exception as e:
        print(f"    ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    # Step 3: Verify output
    print("\n[3/3] Verifying output file...")
    try:
        from openpyxl import load_workbook

        wb = load_workbook(output_path, data_only=False)

        # Check Bill Data sheet
        ws = wb["Bill Data"]
        checks = {
            "Consumer Name": ws["B3"].value,
            "Consumer Number": ws["B4"].value,
            "Units Consumed": ws["B6"].value,
            "Sanctioned Load": ws["B7"].value,
            "Total Bill Amount": ws["B9"].value,
        }

        all_ok = True
        for field, val in checks.items():
            if val is not None and val != 0 and val != "":
                print(f"    ✓ {field}: {val}")
            else:
                print(f"    ✗ {field}: MISSING (got: {val!r})")
                all_ok = False

        # Check Solar Calculation sheet has formulas intact
        ws2 = wb["Solar Calculation"]
        system_size_formula = ws2["B11"].value
        if isinstance(system_size_formula, str) and system_size_formula.startswith("="):
            print(f"    ✓ Formulas intact: System Size = {system_size_formula}")
        else:
            print(f"    ✗ Formula may be missing: B11 = {system_size_formula!r}")
            all_ok = False

        print()
        if all_ok:
            print("✅ ALL CHECKS PASSED")
        else:
            print("⚠️  Some checks failed — see above")

        print(f"\n📊 Output file: {output_path}")
        print("   Open in Excel/Google Sheets to view calculations.\n")

    except Exception as e:
        print(f"    ✗ Verification error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    run_test()
